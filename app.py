"""
ADR 1000-Punkte-Rechner — Flask Application
Gefahrgut-Transportberechnung nach ADR 1.1.3.6 (1000-Punkte-Regel)
"""

import os

from flask import (
    Flask,
    render_template,
    request,
    jsonify,
    redirect,
    url_for,
    abort,
    send_file,
)
from database import get_db, init_db, seed_un_numbers
from befoerderungspapier import generate_befoerderungspapier
from adr_import import parse_adr_pdf, import_adr_data, get_version_history

app = Flask(__name__)
app.config.update(
    # Deutsch-spezifische Einstellungen
    JSON_SORT_KEYS=False,
    JSON_AS_ASCII=False,
    # Flask-spezifische Einstellungen
    SECRET_KEY="adr-1000-punkte-rechner-geheim",  # In Produktion Umgebungsvariable verwenden
    DEBUG=True,
)

# ----------------------------------------------------------------
# Datenbank-Initialisierung beim ersten Start
# ----------------------------------------------------------------
with app.app_context():
    init_db()
    # SQLite WAL-Modus für gleichzeitige Zugriffe (13 Standorte)
    db = get_db()
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA synchronous=NORMAL")
    db.execute("PRAGMA busy_timeout=5000")
    count = db.execute("SELECT COUNT(*) FROM un_numbers").fetchone()[0]
    db.close()
    if count == 0:
        seed_un_numbers()


# ----------------------------------------------------------------
# Hilfsfunktionen
# ----------------------------------------------------------------
def get_db_conn():
    """Wrapper für den Datenbankzugriff."""
    return get_db()


# ----------------------------------------------------------------
# Seiten-Routen
# ----------------------------------------------------------------

@app.route("/")
def index():
    """Hauptseite — 1000-Punkte-Rechner."""
    return render_template("index.html", title="1000-Punkte-Rechner")


@app.route("/befoerderungspapier/<int:id>")
def view_transport_document(id):
    """Beförderungspapier anzeigen — Sendungsdetails und Download-Link."""
    db = get_db()

    shipment = db.execute(
        "SELECT s.*, c.name AS customer_name, sa.name AS sender_name "
        "FROM shipments s "
        "LEFT JOIN customers c ON s.customer_id = c.id "
        "LEFT JOIN shipping_addresses sa ON s.shipping_address_id = sa.id "
        "WHERE s.id = ?",
        (id,),
    ).fetchone()

    if shipment is None:
        db.close()
        abort(404, description=f"Sendung #{id} nicht gefunden.")

    items = db.execute(
        "SELECT * FROM shipment_items WHERE shipment_id = ? ORDER BY id",
        (id,),
    ).fetchall()

    db.close()

    return render_template(
        "befoerderungspapier.html",
        title="Beförderungspapier",
        shipment=shipment,
        items=items,
        shipment_id=id,
    )


@app.route("/befoerderungspapier/<int:id>/pdf")
def download_transport_pdf(id):
    """Beförderungspapier als PDF generieren und zum Download anbieten."""
    try:
        filepath = generate_befoerderungspapier(id)
    except ValueError as e:
        abort(404, description=str(e))
    except Exception as e:
        abort(500, description=f"Fehler bei der PDF-Generierung: {e}")

    filename = os.path.basename(filepath)
    return send_file(
        filepath,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/kunden")
def customers():
    """Kundenverwaltungsseite."""
    return render_template("kunden.html", title="Kundenverwaltung")


@app.route("/adressen")
def addresses():
    """Adressverwaltungsseite."""
    return render_template("adressen.html", title="Adressverwaltung")


@app.route("/un-datenbank")
def un_database():
    """UN-Nummern-Datenbankseite."""
    db = get_db()
    un_count = db.execute("SELECT COUNT(*) FROM un_numbers").fetchone()[0]
    db.close()
    return render_template("un_datenbank.html", title="UN-Datenbank", un_count=un_count)


@app.route("/adr-import")
def adr_import_page():
    """ADR PDF Import-Seite."""
    return render_template("adr_import.html", title="ADR-Version importieren")


# ----------------------------------------------------------------
# API-Routen
# ----------------------------------------------------------------

@app.route("/calculate", methods=["POST"])
def calculate():
    """API-Endpunkt für die 1000-Punkte-Berechnung.

    Erwartet JSON-Daten mit:
        - items: Liste von {un_number, quantity, unit}
        - customer_id: int
        - shipping_address_id: int

    Antwort:
        {
            "total_points": 123.45,
            "is_exempt": true/false,
            "items": [{un_number, substance_name, quantity, unit, category, factor, points}],
            "shipment_id": 1
        }
    """
    data = request.get_json(force=True, silent=True)
    if data is None:
        return jsonify({"error": "Keine gültigen JSON-Daten erhalten"}), 400

    items_in = data.get("items", [])
    customer_id = data.get("customer_id")
    shipping_address_id = data.get("shipping_address_id")

    if not items_in or not isinstance(items_in, list):
        return jsonify({"error": "Keine Gefahrgutpositionen angegeben"}), 400
    if not customer_id or not shipping_address_id:
        return jsonify({"error": "Kunde und Versandadresse sind erforderlich"}), 400

    db = get_db()
    calculated_items = []
    total_points = 0.0

    for item in items_in:
        un_number = str(item.get("un_number", "")).strip()
        quantity = float(item.get("quantity", 0))
        unit = str(item.get("unit", "kg")).strip()
        num_packages = int(item.get("num_packages", 1)) or 1
        package_type = str(item.get("package_type", "")).strip()

        if not un_number or quantity <= 0:
            continue

        # Look up UN number in database
        row = db.execute(
            "SELECT un_number, substance_name_de, hazard_class, transport_category, "
            "points_factor, packing_group FROM un_numbers WHERE un_number = ?",
            (un_number,)
        ).fetchone()

        if row is None:
            continue

        factor = float(row["points_factor"]) if row["points_factor"] is not None else 0.0
        category = row["transport_category"]

        # Category 4 = unlimited → factor is treated as 0 for 1000-point rule
        if category == 4:
            factor = 0.0

        item_points = round(quantity * factor, 2)
        total_points += item_points

        calculated_items.append({
            "un_number": row["un_number"],
            "substance_name": row["substance_name_de"],
            "quantity": quantity,
            "unit": unit,
            "category": category,
            "factor": factor,
            "points": item_points,
            "num_packages": num_packages,
            "package_type": package_type,
        })

    total_points = round(total_points, 2)
    is_exempt = total_points <= 1000

    # Save to shipments and shipment_items tables
    cursor = db.cursor()

    # Determine ADR version — use requested version or latest imported version
    adr_version = data.get("adr_version", "").strip()
    if not adr_version:
        # Use the latest imported ADR version, fallback to 'ADR 2025'
        latest = db.execute(
            "SELECT version FROM adr_versions ORDER BY import_date DESC LIMIT 1"
        ).fetchone()
        adr_version = latest["version"] if latest else "ADR 2025"

    cursor.execute(
        "INSERT INTO shipments (customer_id, shipping_address_id, total_points, is_exempt, adr_version) "
        "VALUES (?, ?, ?, ?, ?)",
        (customer_id, shipping_address_id, total_points, is_exempt, adr_version)
    )
    shipment_id = cursor.lastrowid

    for ci in calculated_items:
        cursor.execute(
            "INSERT INTO shipment_items (shipment_id, un_number, substance_name, "
            "quantity, unit, transport_category, points_factor, item_points, "
            "num_packages, package_type) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (shipment_id, ci["un_number"], ci["substance_name"],
             ci["quantity"], ci["unit"], ci["category"], ci["factor"], ci["points"],
             ci["num_packages"], ci["package_type"])
        )

    db.commit()
    db.close()

    return jsonify({
        "total_points": total_points,
        "is_exempt": is_exempt,
        "items": calculated_items,
        "shipment_id": shipment_id,
    })


# ── ADR PDF Import API ─────────────────────────────────────────────────

@app.route("/api/adr/preview", methods=["POST"])
def adr_preview():
    """Parse ADR PDF without saving — return preview data."""
    if "pdfFile" not in request.files:
        return jsonify({"error": "Keine PDF-Datei hochgeladen"}), 400

    pdf_file = request.files["pdfFile"]
    if not pdf_file.filename:
        return jsonify({"error": "Keine Datei ausgewählt"}), 400

    if not pdf_file.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Nur PDF-Dateien werden unterstützt"}), 400

    version_name = request.form.get("versionName", "ADR 2025").strip()

    try:
        entries = parse_adr_pdf(pdf_file, version_name)
    except Exception as e:
        return jsonify({
            "error": f"Fehler beim Parsen der PDF-Datei: {e}"
        }), 400

    if not entries:
        return jsonify({
            "error": "Keine UN-Nummern in der PDF erkannt. "
                     "Stellen Sie sicher, dass es sich um eine ADR Chapter 3.2 Table A PDF handelt.",
            "entries": [],
        }), 200

    return jsonify({
        "entries": entries,
        "count": len(entries),
        "version": version_name,
    })


@app.route("/api/adr/import", methods=["POST"])
def adr_import():
    """Parse ADR PDF and save to database."""
    if "pdfFile" not in request.files:
        return jsonify({"error": "Keine PDF-Datei hochgeladen"}), 400

    pdf_file = request.files["pdfFile"]
    if not pdf_file.filename:
        return jsonify({"error": "Keine Datei ausgewählt"}), 400

    if not pdf_file.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Nur PDF-Dateien werden unterstützt"}), 400

    version_name = request.form.get("versionName", "ADR 2025").strip()

    try:
        entries = parse_adr_pdf(pdf_file, version_name)
    except Exception as e:
        return jsonify({
            "error": f"Fehler beim Parsen der PDF-Datei: {e}"
        }), 400

    if not entries:
        return jsonify({
            "error": "Keine UN-Nummern in der PDF erkannt. "
                     "Import wurde abgebrochen.",
        }), 400

    # Save to database
    file_path = pdf_file.filename
    result = import_adr_data(entries, version_name, file_path)

    return jsonify(result), 201


@app.route("/api/adr/versions", methods=["GET"])
def adr_versions():
    """List ADR version import history."""
    try:
        history = get_version_history()
        return jsonify(history)
    except Exception as e:
        return jsonify({"error": f"Fehler beim Abrufen der Versionen: {e}"}), 500


@app.route("/api/un-search", methods=["GET"])
def un_search():
    """UN-Nummern-Suche.

    Query-Parameter:
        q (str): Suchbegriff (UN-Nummer oder Stoffname)

    Antwort:
        Liste von UN-Nummern-Datensätzen, die dem Suchbegriff entsprechen.
        [{id, un_number, substance_name_de, substance_name_en, hazard_class,
          transport_category, points_factor, packing_group, tunnel_code,
          special_provisions, max_quantity_per_transport, adr_version}]
    """
    query = request.args.get("q", "").strip()
    if not query:
        return jsonify([])

    db = get_db()
    search_term = f"%{query}%"
    rows = db.execute(
        "SELECT id, un_number, substance_name_de, substance_name_en, hazard_class, "
        "transport_category, points_factor, packing_group, tunnel_code, "
        "special_provisions, max_quantity_per_transport, adr_version "
        "FROM un_numbers "
        "WHERE un_number LIKE ? OR substance_name_de LIKE ? "
        "ORDER BY un_number ASC "
        "LIMIT 50",
        (search_term, search_term)
    ).fetchall()
    db.close()

    results = []
    for row in rows:
        factor = float(row["points_factor"]) if row["points_factor"] is not None else 0.0
        if row["transport_category"] == 4:
            factor = 0.0
        results.append({
            "id": row["id"],
            "un_number": row["un_number"],
            "substance_name_de": row["substance_name_de"],
            "substance_name_en": row["substance_name_en"],
            "hazard_class": row["hazard_class"],
            "transport_category": row["transport_category"],
            "points_factor": factor,
            "packing_group": row["packing_group"],
            "tunnel_code": row["tunnel_code"],
            "special_provisions": row["special_provisions"],
            "max_quantity_per_transport": row["max_quantity_per_transport"],
            "adr_version": row["adr_version"],
        })

    return jsonify(results)


# ── UN Database CRUD API ─────────────────────────────────────────────────

@app.route("/api/un-database", methods=["GET", "POST"])
def api_un_database():
    """UN-Nummern-Datenbank — paginierte Liste & neue Einträge.

    GET  — Paginierte Liste mit Filtern:
        ?q=        (Suchbegriff: UN-Nummer oder Stoffname)
        ?category= (0-4)
        ?class=    (Gefahrklasse)
        ?page=1&per_page=50
        Rückgabe: {items: [...], total: N, page: N, pages: N}

    POST — Neue UN-Nummer anlegen (JSON).
    """
    if request.method == "GET":
        return _un_database_list()
    elif request.method == "POST":
        return _un_database_create()


def _un_database_list():
    """Build paginated, filtered UN number list."""
    q = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    hazard_class = request.args.get("class", "").strip()
    try:
        page = int(request.args.get("page", 1))
        per_page = int(request.args.get("per_page", 50))
    except (ValueError, TypeError):
        page = 1
        per_page = 50

    if page < 1:
        page = 1
    if per_page < 1 or per_page > 200:
        per_page = 50

    db = get_db()

    # Build WHERE clause
    conditions = []
    params = []

    if q:
        search = f"%{q}%"
        conditions.append("(un_number LIKE ? OR substance_name_de LIKE ?)")
        params.extend([search, search])

    if category and category in ("0", "1", "2", "3", "4"):
        conditions.append("transport_category = ?")
        params.append(int(category))

    if hazard_class:
        conditions.append("hazard_class = ?")
        params.append(hazard_class)

    where_clause = ""
    if conditions:
        where_clause = "WHERE " + " AND ".join(conditions)

    # Count total
    count_row = db.execute(
        f"SELECT COUNT(*) AS cnt FROM un_numbers {where_clause}",
        params
    ).fetchone()
    total = count_row["cnt"] if count_row else 0

    total_pages = max(1, (total + per_page - 1) // per_page)
    if page > total_pages:
        page = total_pages

    offset = (page - 1) * per_page

    # Fetch page
    rows = db.execute(
        f"SELECT id, un_number, substance_name_de, substance_name_en, hazard_class, "
        f"packing_group, transport_category, tunnel_code, special_provisions, "
        f"points_factor, max_quantity_per_transport, adr_version "
        f"FROM un_numbers {where_clause} "
        f"ORDER BY un_number ASC "
        f"LIMIT ? OFFSET ?",
        params + [per_page, offset]
    ).fetchall()
    db.close()

    items = []
    for row in rows:
        items.append({
            "id": row["id"],
            "un_number": row["un_number"],
            "substance_name_de": row["substance_name_de"],
            "substance_name_en": row["substance_name_en"],
            "hazard_class": row["hazard_class"],
            "packing_group": row["packing_group"],
            "transport_category": row["transport_category"],
            "tunnel_code": row["tunnel_code"],
            "special_provisions": row["special_provisions"],
            "points_factor": float(row["points_factor"]) if row["points_factor"] is not None else None,
            "max_quantity_per_transport": float(row["max_quantity_per_transport"]) if row["max_quantity_per_transport"] is not None else None,
            "adr_version": row["adr_version"],
        })

    return jsonify({
        "items": items,
        "total": total,
        "page": page,
        "pages": total_pages,
    })


def _un_database_create():
    """Create a new UN number entry."""
    data = request.get_json(force=True, silent=True)
    if data is None:
        return jsonify({"error": "Keine gültigen JSON-Daten erhalten"}), 400

    un_number = str(data.get("un_number", "")).strip()
    substance_name_de = (data.get("substance_name_de") or "").strip()

    if not un_number:
        return jsonify({"error": "UN-Nummer ist erforderlich"}), 400
    if not substance_name_de:
        return jsonify({"error": "Stoffbezeichnung (DE) ist erforderlich"}), 400

    hazard_class = data.get("hazard_class")
    transport_category = data.get("transport_category")

    if not hazard_class:
        return jsonify({"error": "Gefahrklasse ist erforderlich"}), 400
    if transport_category is None:
        return jsonify({"error": "Beförderungskategorie ist erforderlich"}), 400

    db = get_db()

    # Check uniqueness
    existing = db.execute(
        "SELECT id FROM un_numbers WHERE un_number = ?", (un_number,)
    ).fetchone()
    if existing:
        db.close()
        return jsonify({"error": f"UN-Nummer {un_number} existiert bereits"}), 409

    cursor = db.cursor()
    cursor.execute(
        "INSERT INTO un_numbers "
        "(un_number, substance_name_de, substance_name_en, hazard_class, packing_group, "
        " transport_category, points_factor, tunnel_code, special_provisions, "
        " max_quantity_per_transport, adr_version) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (
            un_number,
            substance_name_de,
            (data.get("substance_name_en") or "").strip() or None,
            hazard_class,
            data.get("packing_group") or None,
            int(transport_category),
            float(data.get("points_factor", 0)) if data.get("points_factor") is not None else None,
            (data.get("tunnel_code") or "").strip() or None,
            (data.get("special_provisions") or "").strip() or None,
            float(data.get("max_quantity_per_transport")) if data.get("max_quantity_per_transport") is not None else None,
            (data.get("adr_version") or "ADR 2025").strip(),
        )
    )
    db.commit()
    new_id = cursor.lastrowid

    row = db.execute("SELECT * FROM un_numbers WHERE id = ?", (new_id,)).fetchone()
    db.close()
    return jsonify(dict(row)), 201


@app.route("/api/un-database/<int:id>", methods=["PUT"])
def api_un_database_update(id):
    """UN-Nummern-Eintrag aktualisieren.

    Aktualisierbare Felder: transport_category, points_factor, tunnel_code,
    special_provisions (sowie alle Felder bei vollständigem Update).
    """
    db = get_db()

    row = db.execute("SELECT id FROM un_numbers WHERE id = ?", (id,)).fetchone()
    if row is None:
        db.close()
        return jsonify({"error": f"UN-Nummer #{id} nicht gefunden"}), 404

    data = request.get_json(force=True, silent=True)
    if data is None:
        db.close()
        return jsonify({"error": "Keine gültigen JSON-Daten erhalten"}), 400

    # Build dynamic UPDATE — only set fields that are provided
    allowed_fields = [
        "un_number", "substance_name_de", "substance_name_en", "hazard_class",
        "packing_group", "transport_category", "points_factor", "tunnel_code",
        "special_provisions", "max_quantity_per_transport", "adr_version",
    ]

    set_clauses = []
    params = []

    for field in allowed_fields:
        if field in data:
            if field == "transport_category":
                set_clauses.append("transport_category = ?")
                params.append(int(data[field]))
            elif field in ("points_factor", "max_quantity_per_transport"):
                set_clauses.append(f"{field} = ?")
                val = data[field]
                params.append(float(val) if val is not None and val != "" else None)
            else:
                set_clauses.append(f"{field} = ?")
                val = data[field]
                params.append(str(val).strip() if val else None)

    if not set_clauses:
        db.close()
        return jsonify({"error": "Keine Felder zum Aktualisieren angegeben"}), 400

    set_clauses.append("updated_at = CURRENT_TIMESTAMP")
    params.append(id)

    cursor = db.cursor()
    cursor.execute(
        f"UPDATE un_numbers SET {', '.join(set_clauses)} WHERE id = ?",
        params
    )
    db.commit()

    updated = db.execute("SELECT * FROM un_numbers WHERE id = ?", (id,)).fetchone()
    db.close()
    return jsonify(dict(updated))


@app.route("/api/un-database/stats", methods=["GET"])
def api_un_database_stats():
    """Statistik für die UN-Datenbank.

    Rückgabe:
        {
            "total_count": N,
            "category_distribution": {0: N, 1: N, 2: N, 3: N, 4: N},
            "class_distribution": {"2.2": N, "3": N, ...}
        }
    """
    db = get_db()

    total = db.execute("SELECT COUNT(*) AS cnt FROM un_numbers").fetchone()["cnt"]

    cat_rows = db.execute(
        "SELECT transport_category, COUNT(*) AS cnt "
        "FROM un_numbers GROUP BY transport_category "
        "ORDER BY transport_category"
    ).fetchall()

    category_distribution = {0: 0, 1: 0, 2: 0, 3: 0, 4: 0}
    for row in cat_rows:
        if row["transport_category"] is not None:
            category_distribution[int(row["transport_category"])] = row["cnt"]

    class_rows = db.execute(
        "SELECT hazard_class, COUNT(*) AS cnt "
        "FROM un_numbers GROUP BY hazard_class "
        "ORDER BY hazard_class"
    ).fetchall()

    class_distribution = {}
    for row in class_rows:
        if row["hazard_class"]:
            class_distribution[row["hazard_class"]] = row["cnt"]

    db.close()

    return jsonify({
        "total_count": total,
        "category_distribution": category_distribution,
        "class_distribution": class_distribution,
    })


@app.route("/api/kunden", methods=["GET", "POST"])
def api_customers():
    """Kunden-CRUD-API.

    GET  — Kundenliste abrufen (optional: ?q=Suchbegriff)
    POST — Neuen Kunden anlegen (JSON) ODER Excel-Datei importieren (multipart)
    """
    if request.method == "GET":
        query = request.args.get("q", "").strip()
        db = get_db()
        if query:
            search = f"%{query}%"
            rows = db.execute(
                "SELECT id, name, street, zip, city, country, contact, phone, email "
                "FROM customers "
                "WHERE name LIKE ? OR city LIKE ? OR street LIKE ? OR zip LIKE ? "
                "ORDER BY name ASC LIMIT 200",
                (search, search, search, search)
            ).fetchall()
        else:
            rows = db.execute(
                "SELECT id, name, street, zip, city, country, contact, phone, email "
                "FROM customers "
                "ORDER BY name ASC LIMIT 500"
            ).fetchall()
        db.close()
        customers = [dict(row) for row in rows]
        return jsonify(customers)

    elif request.method == "POST":
        # Check if this is a file upload (Excel import)
        if request.content_type and "multipart" in request.content_type:
            return _handle_excel_import()

        # JSON create
        data = request.get_json(force=True, silent=True)
        if data is None:
            return jsonify({"error": "Keine gültigen JSON-Daten erhalten"}), 400

        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"error": "Der Name ist erforderlich"}), 400

        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            "INSERT INTO customers (name, street, zip, city, country, contact, phone, email) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (
                name,
                (data.get("street") or "").strip(),
                (data.get("zip") or "").strip(),
                (data.get("city") or "").strip(),
                (data.get("country") or "Deutschland").strip(),
                (data.get("contact") or "").strip(),
                (data.get("phone") or "").strip(),
                (data.get("email") or "").strip(),
            )
        )
        db.commit()
        new_id = cursor.lastrowid

        # Return the full record
        row = db.execute("SELECT * FROM customers WHERE id = ?", (new_id,)).fetchone()
        db.close()
        return jsonify(dict(row)), 201


@app.route("/api/kunden/<int:id>", methods=["GET", "PUT", "DELETE"])
def api_customer_by_id(id):
    """Kunden-CRUD für einen einzelnen Kunden.

    GET    — Kundendetails abrufen
    PUT    — Kundendaten aktualisieren
    DELETE — Kunden löschen
    """
    db = get_db()

    if request.method == "GET":
        row = db.execute("SELECT * FROM customers WHERE id = ?", (id,)).fetchone()
        db.close()
        if row is None:
            return jsonify({"error": f"Kunde #{id} nicht gefunden"}), 404
        return jsonify(dict(row))

    elif request.method == "PUT":
        row = db.execute("SELECT id FROM customers WHERE id = ?", (id,)).fetchone()
        if row is None:
            db.close()
            return jsonify({"error": f"Kunde #{id} nicht gefunden"}), 404

        data = request.get_json(force=True, silent=True)
        if data is None:
            db.close()
            return jsonify({"error": "Keine gültigen JSON-Daten erhalten"}), 400

        name = (data.get("name") or "").strip()
        if not name:
            db.close()
            return jsonify({"error": "Der Name ist erforderlich"}), 400

        cursor = db.cursor()
        cursor.execute(
            "UPDATE customers "
            "SET name = ?, street = ?, zip = ?, city = ?, country = ?, "
            "    contact = ?, phone = ?, email = ? "
            "WHERE id = ?",
            (
                name,
                (data.get("street") or "").strip(),
                (data.get("zip") or "").strip(),
                (data.get("city") or "").strip(),
                (data.get("country") or "Deutschland").strip(),
                (data.get("contact") or "").strip(),
                (data.get("phone") or "").strip(),
                (data.get("email") or "").strip(),
                id,
            )
        )
        db.commit()

        updated = db.execute("SELECT * FROM customers WHERE id = ?", (id,)).fetchone()
        db.close()
        return jsonify(dict(updated))

    elif request.method == "DELETE":
        row = db.execute("SELECT id, name FROM customers WHERE id = ?", (id,)).fetchone()
        if row is None:
            db.close()
            return jsonify({"error": f"Kunde #{id} nicht gefunden"}), 404

        # Check for related shipments
        shipments = db.execute(
            "SELECT COUNT(*) as cnt FROM shipments WHERE customer_id = ?", (id,)
        ).fetchone()
        if shipments and shipments["cnt"] > 0:
            db.close()
            return jsonify({
                "error": f"Kunde „{row['name']}” kann nicht gelöscht werden: "
                         f"{shipments['cnt']} Sendung(en) vorhanden. "
                         f"Löschen Sie zuerst die zugehörigen Sendungen."
            }), 409

        db.execute("DELETE FROM customers WHERE id = ?", (id,))
        db.commit()
        db.close()
        return jsonify({"message": f"Kunde „{row['name']}” wurde gelöscht", "id": id})


# ── Excel Import ───────────────────────────────────────────────────────

EXCEL_COLUMNS = ["Name", "Straße", "PLZ", "Stadt", "Land",
                 "Ansprechpartner", "Telefon", "E-Mail"]


def _handle_excel_import():
    """Verarbeitet einen Excel-Upload und importiert Kunden."""
    if "file" not in request.files:
        return jsonify({"error": "Keine Datei hochgeladen"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Keine Datei ausgewählt"}), 400

    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Nur .xlsx-Dateien werden unterstützt"}), 400

    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl ist nicht installiert"}), 500

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({"error": f"Excel-Datei konnte nicht gelesen werden: {e}"}), 400

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return jsonify({"error": "Die Excel-Datei muss eine Kopfzeile und mindestens eine Datenzeile enthalten"}), 400

    # Parse header
    header = [str(c).strip() if c else "" for c in rows[0]]
    col_map = {}
    for expected_col in EXCEL_COLUMNS:
        if expected_col in header:
            col_map[expected_col] = header.index(expected_col)

    if "Name" not in col_map:
        return jsonify({
            "error": "Spalte „Name“ nicht gefunden. Erwartete Spalten: "
                     + ", ".join(EXCEL_COLUMNS)
        }), 400

    db = get_db()
    cursor = db.cursor()

    imported = 0
    updated = 0
    errors = []

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            def cell(col_name):
                idx = col_map.get(col_name)
                if idx is not None and idx < len(row):
                    val = row[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            name = cell("Name")
            if not name:
                errors.append(f"Zeile {row_num}: Name fehlt — übersprungen")
                continue

            street = cell("Straße")
            zip_code = cell("PLZ")
            city = cell("Stadt")
            country = cell("Land") or "Deutschland"
            contact = cell("Ansprechpartner")
            phone = cell("Telefon")
            email = cell("E-Mail")

            # Check if customer exists by name
            existing = db.execute(
                "SELECT id FROM customers WHERE name = ?", (name,)
            ).fetchone()

            if existing:
                cursor.execute(
                    "UPDATE customers "
                    "SET street = ?, zip = ?, city = ?, country = ?, "
                    "    contact = ?, phone = ?, email = ? "
                    "WHERE id = ?",
                    (street, zip_code, city, country, contact, phone, email, existing["id"])
                )
                updated += 1
            else:
                cursor.execute(
                    "INSERT INTO customers (name, street, zip, city, country, contact, phone, email) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (name, street, zip_code, city, country, contact, phone, email)
                )
                imported += 1

        except Exception as e:
            errors.append(f"Zeile {row_num}: {e}")

    db.commit()
    db.close()

    return jsonify({
        "imported": imported,
        "updated": updated,
        "errors": errors,
    }), 201


@app.route("/api/kunden/template", methods=["GET"])
def api_customers_template():
    """Excel-Vorlage für Kundenimport herunterladen."""
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kunden"

    # Header row (bold)
    header_font = openpyxl.styles.Font(bold=True)
    for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font

    # Example rows
    example_data = [
        ["Musterfirma GmbH", "Industriestraße 42", "80331", "München", "Deutschland",
         "Herr Schmidt", "+49 89 123456-0", "info@musterfirma.de"],
        ["Chemiehandel AG", "Hafenweg 10", "20457", "Hamburg", "Deutschland",
         "Frau Müller", "+49 40 987654-0", "bestellung@chemiehandel.de"],
    ]

    for row_idx, row_data in enumerate(example_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths
    for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="kunden_vorlage.xlsx",
    )


@app.route("/api/shipping-addresses", methods=["GET", "POST"])
def api_shipping_addresses():
    """Versandadressen-CRUD-API.

    GET  — Alle Versandadressen zurückgeben.
    POST — Neue Versandadresse anlegen.
    """
    if request.method == "GET":
        db = get_db()
        rows = db.execute(
            "SELECT id, name, street, zip, city, country, is_default "
            "FROM shipping_addresses ORDER BY is_default DESC, name ASC"
        ).fetchall()
        db.close()
        addresses = [dict(row) for row in rows]
        return jsonify(addresses)

    elif request.method == "POST":
        data = request.get_json(force=True, silent=True)
        if data is None:
            return jsonify({"error": "Keine gültigen JSON-Daten erhalten"}), 400

        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"error": "Der Name ist erforderlich"}), 400

        is_default = bool(data.get("is_default", False))

        db = get_db()
        cursor = db.cursor()

        # If setting as default, unset all others first
        if is_default:
            cursor.execute("UPDATE shipping_addresses SET is_default = 0")

        cursor.execute(
            "INSERT INTO shipping_addresses (name, street, zip, city, country, is_default) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (
                name,
                (data.get("street") or "").strip(),
                (data.get("zip") or "").strip(),
                (data.get("city") or "").strip(),
                (data.get("country") or "Deutschland").strip(),
                is_default,
            )
        )
        db.commit()
        new_id = cursor.lastrowid

        row = db.execute(
            "SELECT * FROM shipping_addresses WHERE id = ?", (new_id,)
        ).fetchone()
        db.close()
        return jsonify(dict(row)), 201


@app.route("/api/shipping-addresses/<int:id>", methods=["GET", "PUT", "DELETE"])
def api_shipping_address_by_id(id):
    """Versandadressen-CRUD für eine einzelne Adresse.

    GET    — Adressdetails abrufen
    PUT    — Adresse aktualisieren
    DELETE — Adresse löschen
    """
    db = get_db()

    if request.method == "GET":
        row = db.execute(
            "SELECT * FROM shipping_addresses WHERE id = ?", (id,)
        ).fetchone()
        db.close()
        if row is None:
            return jsonify({"error": f"Adresse #{id} nicht gefunden"}), 404
        return jsonify(dict(row))

    elif request.method == "PUT":
        row = db.execute(
            "SELECT id FROM shipping_addresses WHERE id = ?", (id,)
        ).fetchone()
        if row is None:
            db.close()
            return jsonify({"error": f"Adresse #{id} nicht gefunden"}), 404

        data = request.get_json(force=True, silent=True)
        if data is None:
            db.close()
            return jsonify({"error": "Keine gültigen JSON-Daten erhalten"}), 400

        name = (data.get("name") or "").strip()
        if not name:
            db.close()
            return jsonify({"error": "Der Name ist erforderlich"}), 400

        is_default = bool(data.get("is_default", False))

        cursor = db.cursor()

        if is_default:
            cursor.execute("UPDATE shipping_addresses SET is_default = 0")

        cursor.execute(
            "UPDATE shipping_addresses "
            "SET name = ?, street = ?, zip = ?, city = ?, country = ?, is_default = ? "
            "WHERE id = ?",
            (
                name,
                (data.get("street") or "").strip(),
                (data.get("zip") or "").strip(),
                (data.get("city") or "").strip(),
                (data.get("country") or "Deutschland").strip(),
                is_default,
                id,
            )
        )
        db.commit()

        updated = db.execute(
            "SELECT * FROM shipping_addresses WHERE id = ?", (id,)
        ).fetchone()
        db.close()
        return jsonify(dict(updated))

    elif request.method == "DELETE":
        row = db.execute(
            "SELECT id, name FROM shipping_addresses WHERE id = ?", (id,)
        ).fetchone()
        if row is None:
            db.close()
            return jsonify({"error": f"Adresse #{id} nicht gefunden"}), 404

        # Check for related shipments
        shipments = db.execute(
            "SELECT COUNT(*) as cnt FROM shipments WHERE shipping_address_id = ?", (id,)
        ).fetchone()
        if shipments and shipments["cnt"] > 0:
            db.close()
            return jsonify({
                "error": f"Adresse „{row['name']}” kann nicht gelöscht werden: "
                         f"{shipments['cnt']} Sendung(en) vorhanden. "
                         f"Löschen Sie zuerst die zugehörigen Sendungen."
            }), 409

        db.execute("DELETE FROM shipping_addresses WHERE id = ?", (id,))
        db.commit()
        db.close()
        return jsonify({"message": f"Adresse „{row['name']}” wurde gelöscht", "id": id})


@app.route("/api/shipment", methods=["POST"])
def api_create_shipment():
    """Neue Sendung (Beförderungsvorgang) anlegen.

    Erwartet JSON mit:
        - customer_id
        - shipping_address_id
        - items: Liste von {un_number, quantity, unit}

    Antwort:
        Die erstellte Sendung mit berechneten Punkten.
    """
    # TODO: Implementierung
    data = request.get_json(force=True, silent=True)
    if data is None:
        return jsonify({"error": "Keine gültigen JSON-Daten erhalten"}), 400

    return jsonify({
        "shipment": None,
        "message": "Sendungserstellung noch nicht implementiert",
    }), 201


# ----------------------------------------------------------------
# Fehlerbehandlung
# ----------------------------------------------------------------

@app.errorhandler(404)
def not_found(error):
    """404 — Seite nicht gefunden."""
    return render_template("error.html", title="404 — Seite nicht gefunden",
                           error_code=404,
                           error_message="Die angeforderte Seite wurde nicht gefunden."), 404


@app.errorhandler(500)
def internal_error(error):
    """500 — Interner Serverfehler."""
    return render_template("error.html", title="500 — Interner Serverfehler",
                           error_code=500,
                           error_message="Es ist ein interner Fehler aufgetreten. "
                                         "Bitte versuchen Sie es später erneut."), 500


@app.errorhandler(501)
def not_implemented(error):
    """501 — Noch nicht implementiert."""
    return render_template("error.html", title="501 — Nicht implementiert",
                           error_code=501,
                           error_message=str(error)), 501


# ----------------------------------------------------------------
# Anwendungsstart
# ----------------------------------------------------------------

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5050, debug=True)
